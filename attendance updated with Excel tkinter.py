import openpyxl
import tkinter as tk
from tkinter import messagebox

# Function to read from an Excel file
def read_excel(file_path, sheet_name):
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        # Iterate through rows and columns
        for row in sheet.iter_rows(values_only=True):
            for cell_value in row:
                print(cell_value, end="\t")
            print()

        workbook.close()
    except FileNotFoundError:
        print(f"File '{file_path}' not found.")

# Function to write to an Excel file
def write_excel(file_path, sheet_name, data):
    try:
        workbook = openpyxl.load_workbook(file_path)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()

    # Create a new sheet or get an existing one
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    sheet = workbook[sheet_name]

    # Write data to the sheet
    for row_data in data:
        sheet.append(row_data)

    # Save the workbook to the file
    workbook.save(file_path)
    print(f"Data appended to '{file_path}'.")

# List of students
students = ["Raju", "Kamal", "samy", "nesan", "raja"]
current_student_idx = 0

# File path and sheet name
file_path = "example.xlsx"
sheet_name = "Sheet1"

# Function to mark attendance and move to the next student
def mark_attendance():
    global current_student_idx

    attendance = attendance_var.get()
    student_name = students[current_student_idx]

    if attendance == "Present":
        present_students.append(student_name)
    else:
        absent_students.append(student_name)

    current_student_idx += 1

    if current_student_idx < len(students):
        student_label.config(text=f"Mark Attendance for: {students[current_student_idx]}")
    else:
        mark_button.config(state=tk.DISABLED)
        total_students = len(students)
        present_count = len(present_students)
        absent_count = len(absent_students)
        attendance_percentage = (present_count / total_students) * 100
        result_label.config(
            text=f"Attendance Percentage: {attendance_percentage:.2f}%\n"
                 f"Number of students present: {present_count}\n"
                 f"Number of students absent: {absent_count}"
        )
        display_present_absent_lists()

# Function to display present and absent student lists
def display_present_absent_lists():
    present_list.config(
        text="Present students:\n" + "\n".join([f"{i+1}. {student}" for i, student in enumerate(present_students)]),
        justify="left",
    )
    absent_list.config(
        text="Absent students:\n" + "\n".join([f"{i+1}. {student}" for i, student in enumerate(absent_students)]),
        justify="left",
    )

# Create the GUI window
root = tk.Tk()
root.title("Attendance App")

# Styling
root.geometry("400x400")  # Set the initial window size
root.configure(bg="#f0f0f0")  # Background color

attendance_var = tk.StringVar()
attendance_var.set("Present")  # Default value

# Labels and Spacing
student_label = tk.Label(root, text=f"Mark Attendance for: {students[current_student_idx]}", bg="#f0f0f0")
student_label.pack(pady=10, padx=10)  # Add padding

attendance_menu = tk.OptionMenu(root, attendance_var, "Present", "Absent")
attendance_menu.pack(pady=10)  # Add spacing

mark_button = tk.Button(root, text="Mark Attendance", command=mark_attendance, bg="#008CBA", fg="white")  # Set button color
mark_button.pack(pady=10)

result_label = tk.Label(root, text="", bg="#f0f0f0")
result_label.pack(pady=10)

present_students = []
absent_students = []

# Labels to display present and absent students
present_list = tk.Label(root, text="", bg="#f0f0f0")
present_list.pack(pady=10, padx=10)

absent_list = tk.Label(root, text="", bg="#f0f0f0")
absent_list.pack(pady=10, padx=10)

# Main loop
root.mainloop()
